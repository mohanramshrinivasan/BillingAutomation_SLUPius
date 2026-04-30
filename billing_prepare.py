# ==========================================================
# BILLING AUTOMATION SCRIPT - COMMENTED VERSION
# ==========================================================
#
# What this script does in simple words:
#
# 1. Reads billing information from the Excel file.
# 2. Checks only the rows where the Run column says Yes.
# 3. Checks the invoice date columns to decide which email template to use.
# 4. Creates an Outlook draft email for the patron.
# 5. Writes today's date into the correct invoice date column.
# 6. Changes Run from Yes to Done.
# 7. Saves only the changed cells so Excel formatting is protected.
# 8. Creates a log file so you can check what happened during the run.
#
# IMPORTANT FUTURE CHANGES:
# - If your Excel column names change, update input_data.json, not this script.
# - If your template file names change, update input_data.json.
# - If you want to change the date format, update the run_date() function below.
# - If Outlook draft creation changes, update outlook_prepare_windows().
#
# ==========================================================

import sys
import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# ----------------------------------------------------------
# Tee class
# ----------------------------------------------------------
# This class sends printed messages to two places at the same time:
# 1. The terminal/command prompt screen
# 2. The log file
#
# This helps us see what is happening while the script runs and also keeps a saved record.
# ----------------------------------------------------------
class Tee:
    def __init__(self, file_obj):
        self.file_obj = file_obj
        self.original_stdout = sys.stdout

    def write(self, message):
        self.original_stdout.write(message)
        self.file_obj.write(message)

    def flush(self):
        self.original_stdout.flush()
        self.file_obj.flush()


# ----------------------------------------------------------
# Folder and file locations
# ----------------------------------------------------------
# These paths assume your project folder has this structure:
#
# project_folder/
#   billing_automation.py
#   data/
#     input/
#       input_data.json
#     excel/
#       billing_data.xlsx
#     templates/
#       template files here
#   logs/
#
# If you move files to different folders, update these paths.
# ----------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
TEMPLATES_DIR = DATA_DIR / "templates"
INPUT_DIR = DATA_DIR / "input"
EXCEL_DIR = DATA_DIR / "excel"
LOGS_DIR = BASE_DIR / "logs"

CONFIG_FILE = INPUT_DIR / "input_data.json"
EXCEL_FILE = EXCEL_DIR / "billing_data.xlsx"


# ----------------------------------------------------------
# Create timestamp for log file names
# ----------------------------------------------------------
# Example output: 042726-142530
# This means April 27, 2026 at 14:25:30.
# ----------------------------------------------------------
def timestamp() -> str:
    return datetime.now().strftime("%m%d%y-%H%M%S")


# ----------------------------------------------------------
# Date format written into Excel
# ----------------------------------------------------------
# Current setting: mmddyyyy
# Example: April 27, 2026 becomes 04272026
#
# If you want 04/27/2026 instead, change it to:
# return datetime.now().strftime("%m/%d/%Y")
# ----------------------------------------------------------
def run_date() -> str:
    return datetime.now().strftime("%m%d%Y")


# ----------------------------------------------------------
# Make sure a folder exists
# ----------------------------------------------------------
# If the folder does not exist, this creates it.
# ----------------------------------------------------------
def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


# ----------------------------------------------------------
# Load the JSON settings file
# ----------------------------------------------------------
# input_data.json contains column names, template names, subject lines, etc.
# ----------------------------------------------------------
def load_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ----------------------------------------------------------
# Load an email template file
# ----------------------------------------------------------
# Templates are plain text files stored inside data/templates.
# ----------------------------------------------------------
def load_template(path: Path) -> str:
    with open(path, "r", encoding="utf-8") as f:
        return f.read()


# ----------------------------------------------------------
# Check whether a cell has a real value
# ----------------------------------------------------------
# This is used for invoice date fields.
# The script does NOT care what exact date format is inside the cell.
# It only checks whether the cell is blank or filled.
#
# Blank = use the current template stage.
# Filled = move to the next template stage.
# ----------------------------------------------------------
def filled(value) -> bool:
    if value is None:
        return False

    value = str(value).strip()

    if value == "":
        return False

    if value.lower() in ["nan", "nat", "none", "<na>"]:
        return False

    return True


# ----------------------------------------------------------
# Clean text from Excel
# ----------------------------------------------------------
# This prevents ugly values like nan or None from appearing in emails.
# ----------------------------------------------------------
def clean_text(value) -> str:
    if value is None:
        return ""

    value = str(value).strip()

    if value.lower() in ["nan", "nat", "none", "<na>"]:
        return ""

    return value


# ----------------------------------------------------------
# Format patron name
# ----------------------------------------------------------
# If Excel has name like: TABERNER, MOLLY
# This changes it to: Molly Taberner
# ----------------------------------------------------------
def format_name(name) -> str:
    name = clean_text(name)

    if not name:
        return ""

    if "," in name:
        last, first = name.split(",", 1)
        return f"{first.strip().title()} {last.strip().title()}"

    return name.title()


# ----------------------------------------------------------
# Format money for email display
# ----------------------------------------------------------
# Example: 120 becomes 120.00
# The dollar sign is added later in the email text.
# ----------------------------------------------------------
def format_money(value) -> str:
    try:
        return f"{float(value):.2f}"
    except Exception:
        return "0.00"


# ----------------------------------------------------------
# Convert money to number for total calculation
# ----------------------------------------------------------
def money_float(value) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


# ----------------------------------------------------------
# Create safe file names if needed in the future
# ----------------------------------------------------------
# This removes special characters that can break file names.
# ----------------------------------------------------------
def safe_filename(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", str(text)).strip("_")


# ----------------------------------------------------------
# Validate Excel columns
# ----------------------------------------------------------
# This checks whether the column names listed in input_data.json actually exist in Excel.
#
# This is very important because if the config says Invoice_Date1 but Excel says
# Invoice Date 1, the script will not know they are supposed to be the same.
#
# If a column is missing, the script stops and prints the exact problem.
# ----------------------------------------------------------
def validate_columns(df: pd.DataFrame, config: dict, run_column: str) -> None:
    required_columns = [
        run_column,
        config["invoice_date_1_column"],
        config["invoice_date_2_column"],
        config["invoice_date_3_column"],
        config["invoice_date_4_column"],
        config["to_column"],
        config["title_column"],
        config["barcode_column"],
        config["call_number_column"],
        config["fines_column"],
        config["name_column"],
    ]

    missing_columns = [col for col in required_columns if col not in df.columns]

    if missing_columns:
        print("\nERROR: These column names from input_data.json were NOT found in Excel:")
        for col in missing_columns:
            print(f" - {col}")

        print("\nActual Excel columns found:")
        for col in df.columns:
            print(f" - {col}")

        raise ValueError("Column name mismatch. Fix input_data.json or Excel headers.")


# ----------------------------------------------------------
# Decide which invoice/template stage should be used
# ----------------------------------------------------------
# Template logic:
#
# If Invoice_Date1 is empty -> use Template 1
# If Invoice_Date1 is filled and Invoice_Date2 is empty -> use Template 2
# If Invoice_Date1 and Invoice_Date2 are filled and Invoice_Date3 is empty -> use Template 3
# If Invoice_Date1, Invoice_Date2, and Invoice_Date3 are filled and Final date is empty -> use Template 4
# If all invoice dates are filled -> do not create another draft
# ----------------------------------------------------------
def get_stage(row: pd.Series, config: dict) -> int:
    col1 = config["invoice_date_1_column"]
    col2 = config["invoice_date_2_column"]
    col3 = config["invoice_date_3_column"]
    col4 = config["invoice_date_4_column"]

    d1 = row[col1]
    d2 = row[col2]
    d3 = row[col3]
    d4 = row[col4]

    # Debug message helps you confirm what the script is reading from Excel.
    print(
        f"DEBUG Invoice Dates: "
        f"{col1}='{d1}', "
        f"{col2}='{d2}', "
        f"{col3}='{d3}', "
        f"{col4}='{d4}'"
    )

    if not filled(d1):
        return 1

    if not filled(d2):
        return 2

    if not filled(d3):
        return 3

    if not filled(d4):
        return 4

    return 0


# ----------------------------------------------------------
# Get the invoice date column that must be updated after draft creation
# ----------------------------------------------------------
# If Template 1 is used, update Invoice_Date1.
# If Template 2 is used, update Invoice_Date2.
# If Template 3 is used, update Invoice_Date3.
# If Template 4 is used, update Final Invoice Date.
# ----------------------------------------------------------
def get_stage_column(stage: int, config: dict) -> str:
    if stage == 1:
        return config["invoice_date_1_column"]

    if stage == 2:
        return config["invoice_date_2_column"]

    if stage == 3:
        return config["invoice_date_3_column"]

    if stage == 4:
        return config["invoice_date_4_column"]

    return ""


# ----------------------------------------------------------
# Build the email details for one patron
# ----------------------------------------------------------
# If one patron has multiple items, this combines the items into one email.
#
# The returned values are used inside the email template.
# Example template placeholder: {Preferred_Name}
# ----------------------------------------------------------
def build_group_context(group_df: pd.DataFrame, config: dict) -> dict:
    first_row = group_df.iloc[0]

    item_lines_list = []
    item_details_list = []
    total_fines = 0.0

    for _, row in group_df.iterrows():
        title = clean_text(row[config["title_column"]])
        barcode = clean_text(row[config["barcode_column"]])
        call_number = clean_text(row[config["call_number_column"]])
        fines_value = money_float(row[config["fines_column"]])

        total_fines += fines_value

        # Simple one-line item summary
        item_lines_list.append(f"{title} - ${format_money(fines_value)}")

        # Detailed item information for the email body
        item_details_list.append(
            f"Title/Author: {title}\n"
            f"Call Number: {call_number}\n"
            f"Barcode: {barcode}\n"
            f"Replacement Cost: ${format_money(fines_value)}"
        )

    return {
        "Preferred_Name": format_name(first_row[config["name_column"]]),
        "Title": clean_text(first_row[config["title_column"]]),
        "Barcode": clean_text(first_row[config["barcode_column"]]),
        "Call_Number": clean_text(first_row[config["call_number_column"]]),
        "Fines": format_money(first_row[config["fines_column"]]),
        "Item_Lines": "\n".join(item_lines_list),
        "Item_Details": "\n\n".join(item_details_list),
        "Total_Fines": format_money(total_fines),
        "From_Email": clean_text(config.get("from_email", "")),
        "Today": run_date(),
    }


# ----------------------------------------------------------
# Create Outlook draft on Windows
# ----------------------------------------------------------
# This creates the draft, marks it High Importance, saves it, and opens it.
#
# Important:
# - This works with Outlook Classic desktop app.
# - It may not work with New Outlook because New Outlook has limited COM support.
# ----------------------------------------------------------
def outlook_prepare_windows(to_email: str, subject: str, body: str, from_email: str = "") -> str:
    import win32com.client

    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)

    mail.To = to_email
    mail.Subject = subject
    mail.Body = body

    # Outlook importance values:
    # 0 = Low
    # 1 = Normal
    # 2 = High
    mail.Importance = 2

    # Try to select the correct Outlook sending account.
    # Example: piuscirc@slu.edu
    if from_email:
        session = outlook.Session

        for account in session.Accounts:
            try:
                if str(account.SmtpAddress).lower() == from_email.lower():
                    mail.SendUsingAccount = account
                    break
            except Exception:
                pass

        # This helps with shared mailbox or sending on behalf of another email.
        try:
            mail.SentOnBehalfOfName = from_email
        except Exception:
            pass

    mail.Save()
    mail.Display()

    return "Outlook draft prepared, displayed, and marked High Importance"


# ----------------------------------------------------------
# Save only changed cells back to Excel
# ----------------------------------------------------------
# This is safer than replacing the whole sheet.
# It protects formatting, leading zeros, formulas, colors, and other columns.
#
# It updates only:
# 1. The invoice date column for the current stage
# 2. The Run column
# ----------------------------------------------------------
def save_only_changed_cells_to_excel(
    excel_file: Path,
    sheet_name: str,
    run_column: str,
    changed_rows: dict
) -> None:
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # Create a map of header names to Excel column numbers.
    # Example: Invoice_Date1 -> column 12
    header_map = {}

    for cell in ws[1]:
        if cell.value:
            header_map[str(cell.value).strip()] = cell.column

    if run_column not in header_map:
        raise ValueError(f"Run column not found in Excel: {run_column}")

    for idx, update_info in changed_rows.items():
        # pandas row index starts at 0.
        # Excel row 1 is the header.
        # So data row starts at Excel row 2.
        excel_row = idx + 2

        stage_column = update_info["stage_column"]
        today_value = update_info["today_value"]

        if stage_column not in header_map:
            raise ValueError(f"Invoice date column not found in Excel: {stage_column}")

        # Update the correct invoice date cell.
        ws.cell(row=excel_row, column=header_map[stage_column]).value = today_value

        # Change Run from Yes to Done after successful draft creation.
        ws.cell(row=excel_row, column=header_map[run_column]).value = "Done"

    wb.save(excel_file)


# ----------------------------------------------------------
# Main script
# ----------------------------------------------------------
# This is the main workflow:
# 1. Load settings
# 2. Load templates
# 3. Load Excel
# 4. Find rows marked Run = Yes
# 5. Group rows by email
# 6. Create Outlook drafts
# 7. Update Excel
# ----------------------------------------------------------
def main():
    ensure_dir(LOGS_DIR)

    if not CONFIG_FILE.exists():
        raise FileNotFoundError(f"Missing config file: {CONFIG_FILE}")

    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f"Missing Excel file: {EXCEL_FILE}")

    config = load_json(CONFIG_FILE)

    # Remove accidental spaces from JSON values.
    for key, value in config.items():
        if isinstance(value, str):
            config[key] = value.strip()

    run_stamp = timestamp()
    log_file = LOGS_DIR / f"logs-{run_stamp}.txt"

    log_f = open(log_file, "w", encoding="utf-8")
    sys.stdout = Tee(log_f)
    sys.stderr = Tee(log_f)

    try:
        print(f"Run started: {run_stamp}")
        print(f"Workbook: {EXCEL_FILE}")

        # Load subject lines and template bodies.
        # If you want to change template files or subjects, update input_data.json.
        template_map = {
            1: {
                "subject": config["template_1_subject"],
                "body": load_template(TEMPLATES_DIR / config["template_1_file"]),
            },
            2: {
                "subject": config["template_2_subject"],
                "body": load_template(TEMPLATES_DIR / config["template_2_file"]),
            },
            3: {
                "subject": config["template_3_subject"],
                "body": load_template(TEMPLATES_DIR / config["template_3_file"]),
            },
            4: {
                "subject": config["template_4_subject"],
                "body": load_template(TEMPLATES_DIR / config["template_4_file"]),
            },
        }

        # Read Excel as text.
        # This protects Banner numbers, barcodes, and invoice dates.
        df = pd.read_excel(
            EXCEL_FILE,
            sheet_name=config["sheet_name"],
            dtype=str,
            keep_default_na=False
        )

        # Remove spaces from Excel header names.
        # Example: "Invoice_Date1 " becomes "Invoice_Date1"
        df.columns = df.columns.str.strip()

        if len(df.columns) < 2:
            raise ValueError("Excel must have at least 2 columns. Column 2 should be Run.")

        # The script uses the second Excel column as the Run column.
        # Example: if column B is Run, that is the one used here.
        run_column = df.columns[1]

        # Stop early if required columns are missing.
        validate_columns(df, config, run_column)

        processed_count = 0
        skip_count = 0
        today_value = run_date()
        changed_rows = {}

        # Only rows with Run = Yes will be processed.
        # Run = Done or blank rows will be ignored.
        eligible_df = df[
            df[run_column].fillna("").astype(str).str.strip().str.lower() == "yes"
        ].copy()

        if eligible_df.empty:
            print("No rows found with Run = Yes.")
            print("Please change the Run column to Yes before running the script.")
            return

        # Group rows by email so one patron can receive one combined email.
        grouped = eligible_df.groupby(config["to_column"], dropna=False)

        for to_email, group_df in grouped:
            to_email = clean_text(to_email)

            if not to_email:
                print("Skipped one group because TO email is blank.")
                skip_count += len(group_df)
                continue

            # Check invoice stage for each row in this email group.
            stages = group_df.apply(lambda row: get_stage(row, config), axis=1).tolist()
            unique_stages = set(stages)

            print(f"DEBUG {to_email}: detected stages = {sorted(unique_stages)}")

            # If every invoice date field is filled, there is nothing else to send.
            if unique_stages == {0}:
                print(
                    f"Email {to_email}: skipped because all invoice date fields are already filled."
                )
                skip_count += len(group_df)
                continue

            valid_stages = {s for s in unique_stages if s != 0}

            # Safety rule:
            # If the same patron has multiple rows, all rows must be in the same invoice stage.
            # This prevents combining Template 1 and Template 2 items into one email.
            if len(valid_stages) != 1:
                print(
                    f"Email {to_email}: skipped because rows have mixed invoice stages "
                    f"{sorted(unique_stages)}. Please check invoice date columns."
                )
                skip_count += len(group_df)
                continue

            stage = valid_stages.pop()
            stage_column = get_stage_column(stage, config)

            template_info = template_map[stage]
            subject = template_info["subject"]
            from_email = clean_text(config.get("from_email", ""))

            context = build_group_context(group_df, config)

            try:
                body = template_info["body"].format(**context)
            except KeyError as e:
                print(f"Email {to_email}: template placeholder missing: {e}")
                skip_count += len(group_df)
                continue

            try:
                action_result = outlook_prepare_windows(
                    to_email=to_email,
                    subject=subject,
                    body=body,
                    from_email=from_email
                )
            except Exception as e:
                print(f"Email {to_email}: Outlook draft preparation failed: {e}")
                skip_count += len(group_df)
                continue

            # If draft was created successfully, prepare Excel updates.
            for idx in group_df.index:
                changed_rows[idx] = {
                    "stage_column": stage_column,
                    "today_value": today_value
                }

            processed_count += len(group_df)

            print(
                f"Email {to_email}: {action_result} | "
                f"template stage={stage} | "
                f"updated {stage_column} to {today_value} | "
                f"updated {run_column} to Done | "
                f"rows combined={len(group_df)} | "
                f"total=${context['Total_Fines']}"
            )

        # Write updates back to Excel after all drafts are created.
        save_only_changed_cells_to_excel(
            excel_file=EXCEL_FILE,
            sheet_name=config["sheet_name"],
            run_column=run_column,
            changed_rows=changed_rows
        )

        print(f"Run completed. Processed rows: {processed_count}. Rows skipped: {skip_count}.")
        print(f"Log file: {log_file}")

    finally:
        sys.stdout = sys.__stdout__
        sys.stderr = sys.__stderr__
        log_f.close()


# ----------------------------------------------------------
# Script start point
# ----------------------------------------------------------
# Python starts running the script from here.
# Do not remove this part.
# ----------------------------------------------------------
if __name__ == "__main__":
    main()
